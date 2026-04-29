from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
import datetime

# ── Paleta ────────────────────────────────────────────────────────────────────
AZUL_OSCURO   = "0D2137"
AZUL_MED      = "1A4A7A"
AZUL_CLARO    = "2E86C1"
AZUL_SUAVE    = "D6EAF8"
GRIS_CLARO    = "F2F3F4"
VERDE_OK      = "1E8449"
NARANJA_WARN  = "CA6F1E"
ROJO_BAD      = "C0392B"
BLANCO        = "FFFFFF"
GOLD          = "F4D03F"

# ── Helpers ───────────────────────────────────────────────────────────────────

def _thin(color="BBBBBB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom(color=AZUL_OSCURO):
    thick = Side(style="medium", color=color)
    thin  = Side(style="thin",   color="BBBBBB")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def cel(ws, ref, value=None, bold=False, size=10, color="000000",
        bg=None, align="center", border=False, fmt=None, italic=False,
        wrap=True):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name="Calibri", bold=bold, italic=italic,
                  size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    if border:
        c.border = _thin()
    if fmt:
        c.number_format = fmt

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ── Hoja 1: Resumen Comparativo ───────────────────────────────────────────────
#
# Estructura de encabezados (filas 4–6):
#
#   Fila 4 (escenarios, span 2 cols c/u):
#       [Métrica] [Unidad] [   Real   ] [  Eficiente  ] [ No eficiente ]
#
#   Fila 5 (tipos de cargador):
#       [        ] [      ] [ CR ] [CSR] [ CR ] [ CSR ] [ CR ] [ CSR  ]
#
#   Fila 6: línea separadora visual (altura pequeña, fondo AZUL_CLARO)
#
# Los datos empiezan en fila 7.

def _hoja_resumen(wb, resultados, sim_time, seed, n_replicas):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    # ── Fila 1: Título principal ──────────────────────────────────────────
    ws.merge_cells("A1:J1")
    cel(ws, "A1",
        "ANÁLISIS DE SENSIBILIDAD — ESTACIÓN DE CARGA VE",
        bold=True, size=15, color=BLANCO, bg=AZUL_OSCURO)
    set_row_height(ws, 1, 36)

    # ── Fila 2: Subtítulo con metadatos ───────────────────────────────────
    ws.merge_cells("A2:J2")
    cel(ws, "A2",
        f"  {n_replicas} réplicas  ·  {sim_time} min c/u  ·  "
        f"Seed: {seed or 'aleatorio'}  ·  "
        f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
        italic=True, size=9, color=BLANCO, bg=AZUL_MED, align="left")
    set_row_height(ws, 2, 18)

    # ── Fila 3: banda decorativa ──────────────────────────────────────────
    ws.merge_cells("A3:J3")
    ws["A3"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    set_row_height(ws, 3, 4)

    # ── Fila 4: nombre de cada escenario (span 2 columnas) ────────────────
    # Columnas: A=Métrica, B=Unidad, C–D=Real, E–F=Eficiente, G–H=No eficiente
    set_row_height(ws, 4, 22)

    # Etiquetas fijas
    cel(ws, "A4", "", bg=AZUL_OSCURO)
    cel(ws, "B4", "", bg=AZUL_OSCURO)

    # Escenarios — fondo ligeramente diferente para cada uno
    ESCENARIO_COLS = [("C", "D"), ("E", "F"), ("G", "H")]
    BG_ESCENARIOS  = [AZUL_OSCURO, "163D64", "0A2A47"]   # gradación de azul

    for i, r in enumerate(resultados):
        c_ini, c_fin = ESCENARIO_COLS[i]
        ws.merge_cells(f"{c_ini}4:{c_fin}4")
        cel(ws, f"{c_ini}4", r["nombre"],
            bold=True, size=11, color=GOLD, bg=BG_ESCENARIOS[i])

    # ── Fila 5: tipo de cargador (CR / CSR) ───────────────────────────────
    set_row_height(ws, 5, 20)
    cel(ws, "A5", "Métrica",
        bold=True, size=10, color=BLANCO, bg=AZUL_MED, border=True)
    cel(ws, "B5", "Unidad",
        bold=True, size=10, color=BLANCO, bg=AZUL_MED, border=True)

    BG_CR_CSR = [("1B5E8A", "1A6EA0"),   # Real: CR / CSR
                 ("1A6636", "1E7A3E"),   # Eficiente
                 ("7B2400", "8C2900")]   # No eficiente

    col = 3
    for i in range(len(resultados)):
        for tipo, bg in zip(["CR", "CSR"], BG_CR_CSR[i]):
            ref = f"{get_column_letter(col)}5"
            cel(ws, ref, tipo, bold=True, size=10, color=BLANCO,
                bg=bg, border=True)
            col += 1

    # ── Fila 6: separador ─────────────────────────────────────────────────
    set_row_height(ws, 6, 3)
    for c in range(1, 9):
        ws.cell(row=6, column=c).fill = PatternFill("solid", fgColor=AZUL_CLARO)

    # ── Filas 7+: datos ───────────────────────────────────────────────────
    METRICAS = [
        ("N° Cargadores",        "#",   "n_cargadores",          None,    False),
        ("Llegadas promedio",    "#",   "llegadas",              "0.0",   False),
        ("Atendidos promedio",   "#",   "atendidos",             "0.0",   False),
        ("Arrepentidos prom.",   "#",   "arrepentidos",          "0.0",   False),
        ("% Abandono",           "%",   "pct_arrepentidos",      "0.0%",  True),
        ("Espera promedio",      "min", "tiempo_espera_prom",    "0.00",  False),
        ("Tiempo en sistema",    "min", "tiempo_sistema_prom",   "0.00",  False),
        ("Ociosidad cargadores", "%",   "ociosidad_por_cargador","0.0%",  True),
    ]

    fila = 7
    for idx, (label, unidad, key, fmt, is_pct) in enumerate(METRICAS):
        set_row_height(ws, fila, 20)
        row_bg = GRIS_CLARO if idx % 2 == 0 else BLANCO

        cel(ws, f"A{fila}", label, bold=True, size=10,
            bg=row_bg, align="left", border=True)
        cel(ws, f"B{fila}", unidad, size=9, color="666666",
            bg=row_bg, border=True)

        col = 3
        for r in resultados:
            for tipo in ["CR", "CSR"]:
                s = r[tipo]
                if s["n_cargadores"] == 0:
                    val = "–"
                else:
                    val = s[key]
                    if key == "pct_arrepentidos":
                        val /= 100
                    elif key == "ociosidad_por_cargador":
                        val = (sum(val) / len(val) / 100) if val else 0

                txt_color = "000000"
                if key == "pct_arrepentidos" and val != "–":
                    if val > 0.20:
                        txt_color = ROJO_BAD
                    elif val > 0.10:
                        txt_color = NARANJA_WARN
                    else:
                        txt_color = VERDE_OK

                ref = f"{get_column_letter(col)}{fila}"
                cel(ws, ref, val, size=10, color=txt_color,
                    bg=row_bg, border=True,
                    fmt=fmt if val != "–" else None)
                col += 1
        fila += 1

    # ── Abandono total ────────────────────────────────────────────────────
    set_row_height(ws, fila, 20)
    cel(ws, f"A{fila}", "Abandono total", bold=True, size=10,
        bg=AZUL_SUAVE, border=True)
    cel(ws, f"B{fila}", "%", size=9, color="666666",
        bg=AZUL_SUAVE, border=True)
    col = 3
    for r in resultados:
        total_l = r["CR"]["llegadas"] + r["CSR"]["llegadas"]
        total_a = r["CR"]["arrepentidos"] + r["CSR"]["arrepentidos"]
        val = (total_a / total_l) if total_l > 0 else 0
        for _ in ("CR", "CSR"):
            txt_color = (ROJO_BAD if val > 0.20
                         else (NARANJA_WARN if val > 0.10 else VERDE_OK))
            cel(ws, f"{get_column_letter(col)}{fila}", val,
                size=10, bold=True, color=txt_color,
                bg=AZUL_SUAVE, border=True, fmt="0.0%")
            col += 1
    fila += 1

    # ── Eficiencia global ─────────────────────────────────────────────────
    # NOTA: mayor score = mejor (invertido respecto a versión anterior)
    set_row_height(ws, fila, 24)
    cel(ws, f"A{fila}", "Eficiencia global (⭡ mayor = mejor)",
        bold=True, size=10, bg=AZUL_OSCURO, color=BLANCO, border=True)
    cel(ws, f"B{fila}", "score", size=9, color=GOLD,
        bg=AZUL_OSCURO, border=True)
    col = 3

    # Determinar el mejor score para resaltarlo
    mejor_score = max(r["eficiencia_global"] for r in resultados)

    for r in resultados:
        score = r["eficiencia_global"]
        es_mejor = abs(score - mejor_score) < 1e-9
        # CR column: mostrar score
        cel(ws, f"{get_column_letter(col)}{fila}",
            score,
            size=11, bold=True,
            color=GOLD if es_mejor else "AAAAAA",
            bg=AZUL_OSCURO, border=True, fmt="0.000")
        # Indicador visual si es el mejor
        mejor_txt = "★ MEJOR" if es_mejor else "—"
        cel(ws, f"{get_column_letter(col+1)}{fila}", mejor_txt,
            size=9,
            color=GOLD if es_mejor else "666666",
            bg=AZUL_OSCURO, border=True)
        col += 2

    # ── Anchos de columnas ────────────────────────────────────────────────
    set_col_widths(ws, {
        "A": 26, "B": 8,
        "C": 11, "D": 11,
        "E": 11, "F": 11,
        "G": 11, "H": 11,
        "I": 2,  "J": 2,
    })


# ── Hoja 2: KPIs visuales ─────────────────────────────────────────────────────

def _hoja_kpis(wb, resultados):
    ws = wb.create_sheet("KPIs")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    cel(ws, "A1", "PANEL DE KPIs POR ESCENARIO",
        bold=True, size=14, color=BLANCO, bg=AZUL_OSCURO)
    set_row_height(ws, 1, 32)

    ws.merge_cells("A2:F2")
    cel(ws, "A2", "Comparación rápida entre escenarios simulados",
        italic=True, size=9, color=BLANCO, bg=AZUL_MED)
    set_row_height(ws, 2, 16)

    # KPIs a mostrar, agrupados por tipo
    kpis_cr = [
        ("Abandono total", "pct_arrepentidos", "0.0%"),
        ("Espera prom. CR", "tiempo_espera_prom", "0.00"),
        ("Ociosidad CR", "ociosidad_por_cargador", "0.0%"),
        ("Atendidos CR", "atendidos", "0.0"),
    ]
    kpis_csr = [
        ("Abandono total CSR", "pct_arrepentidos", "0.0%"),
        ("Espera prom. CSR", "tiempo_espera_prom", "0.00"),
        ("Ociosidad CSR", "ociosidad_por_cargador", "0.0%"),
        ("Atendidos CSR", "atendidos", "0.0"),
    ]

    mejor_score = max(r["eficiencia_global"] for r in resultados)

    # Encabezados de escenarios (sin merge para evitar error de MergedCell)
    cel(ws, "A1", "", bg=AZUL_OSCURO)
    cel(ws, "A2", "", bg=AZUL_OSCURO)
    cel(ws, "A3", "", bg=AZUL_OSCURO)
    cel(ws, "A4", "", bg=AZUL_OSCURO)
    for idx, r in enumerate(resultados):
        col = 2 + idx
        es_mejor = abs(r["eficiencia_global"] - mejor_score) < 1e-9
        bg_header = AZUL_CLARO if not es_mejor else "117A65"
        cel(ws, f"{get_column_letter(col)}4", r["nombre"] + (" ★" if es_mejor else ""), bold=True, size=12, color=BLANCO, bg=bg_header, border=True)

    # KPIs CR
    fila = 5
    for label, kkey, kfmt in kpis_cr:
        cel(ws, f"A{fila}", label, bold=True, size=10, color=BLANCO, bg=AZUL_OSCURO, border=True)
        for idx, r in enumerate(resultados):
            col = 2 + idx
            datos = r["CR"]
            if kkey == "pct_arrepentidos":
                val = datos[kkey] / 100
                txt_color = (ROJO_BAD if val > 0.20 else (NARANJA_WARN if val > 0.10 else VERDE_OK))
            elif kkey == "ociosidad_por_cargador":
                oc = datos[kkey]
                val = (sum(oc) / len(oc) / 100) if oc else 0
                txt_color = AZUL_CLARO
            else:
                val = datos.get(kkey, 0)
                txt_color = AZUL_CLARO
            cel(ws, f"{get_column_letter(col)}{fila}", val, bold=True, size=11, color=txt_color, bg=BLANCO, border=True, fmt=kfmt)
        fila += 2

    # KPIs CSR
    for label, kkey, kfmt in kpis_csr:
        cel(ws, f"A{fila}", label, bold=True, size=10, color=BLANCO, bg=AZUL_OSCURO, border=True)
        for idx, r in enumerate(resultados):
            col = 2 + idx
            datos = r["CSR"]
            if kkey == "pct_arrepentidos":
                val = datos[kkey] / 100
                txt_color = (ROJO_BAD if val > 0.20 else (NARANJA_WARN if val > 0.10 else VERDE_OK))
            elif kkey == "ociosidad_por_cargador":
                oc = datos[kkey]
                val = (sum(oc) / len(oc) / 100) if oc else 0
                txt_color = AZUL_CLARO
            else:
                val = datos.get(kkey, 0)
                txt_color = AZUL_CLARO
            cel(ws, f"{get_column_letter(col)}{fila}", val, bold=True, size=11, color=txt_color, bg=BLANCO, border=True, fmt=kfmt)
        fila += 2

    # Separadores visuales y anchos
    for f in range(5, fila, 2):
        for c in range(1, 2+len(resultados)+1):
            ws.cell(row=f+1, column=c).fill = PatternFill("solid", fgColor=AZUL_SUAVE)

    ancho_cols = {"A": 18}
    for i in range(len(resultados)):
        ancho_cols[get_column_letter(2+i)] = 14
    set_col_widths(ws, ancho_cols)


# ── Hoja 3: Insights ──────────────────────────────────────────────────────────

def _hoja_insights(wb, resultados):
    ws = wb.create_sheet("Insights")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    cel(ws, "A1", "INSIGHTS AUTOMÁTICOS",
        bold=True, size=14, color=BLANCO, bg=AZUL_OSCURO)
    set_row_height(ws, 1, 32)

    row = 3
    for r in resultados:
        ws.merge_cells(f"A{row}:E{row}")
        cel(ws, f"A{row}", f"📊  Escenario: {r['nombre']}",
            bold=True, size=12, color=BLANCO, bg=AZUL_MED, align="left")
        set_row_height(ws, row, 24)
        row += 1

        cr = r["CR"]; csr = r["CSR"]
        total_l = cr["llegadas"] + csr["llegadas"]
        total_a = cr["arrepentidos"] + csr["arrepentidos"]
        abandono = total_a / max(1, total_l)

        lines = [
            ("Abandono total:",      f"{abandono:.1%}",
             ROJO_BAD if abandono > 0.20 else (NARANJA_WARN if abandono > 0.10 else VERDE_OK)),
            ("Atendidos CR / CSR:",  f"{cr['atendidos']:.0f} / {csr['atendidos']:.0f}", AZUL_CLARO),
            ("Espera prom. CR:",     f"{cr['tiempo_espera_prom']:.1f} min", "333333"),
            ("Espera prom. CSR:",    f"{csr['tiempo_espera_prom']:.1f} min", "333333"),
            # NOTA: eficiencia_global ahora ↑ mayor = mejor
            ("Eficiencia global:",   f"{r['eficiencia_global']:.3f}  (↑ mayor = mejor)", AZUL_OSCURO),
        ]

        for label, value, vc in lines:
            set_row_height(ws, row, 18)
            cel(ws, f"A{row}", label, bold=True, size=10,
                bg=GRIS_CLARO, align="left", border=True)
            ws.merge_cells(f"B{row}:C{row}")
            cel(ws, f"B{row}", value, bold=True, size=10,
                color=vc, bg=BLANCO, align="left", border=True)
            row += 1

        set_row_height(ws, row, 20)
        ws.merge_cells(f"A{row}:E{row}")
        if abandono > 0.20:
            rec = "⚠️  Alta pérdida de demanda — considerar más cargadores."
            bg_rec = "FADBD8"
        elif abandono > 0.10:
            rec = "⚡  Pérdida moderada — monitorear horas pico."
            bg_rec = "FDEBD0"
        else:
            rec = "✅  Sistema eficiente — demanda bien gestionada."
            bg_rec = "D5F5E3"
        cel(ws, f"A{row}", rec, italic=True, size=10,
            bg=bg_rec, align="left")
        row += 2

    set_col_widths(ws, {"A": 26, "B": 16, "C": 14, "D": 14, "E": 14})


# ── Hoja 4: Gráficos comparativos ────────────────────────────────────────────

def _hoja_graficos(wb, resultados):
    ws = wb.create_sheet("Gráficos")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    cel(ws, "A1", "GRÁFICOS COMPARATIVOS",
        bold=True, size=14, color=BLANCO, bg=AZUL_OSCURO)
    set_row_height(ws, 1, 32)

    headers_g = [
        "Escenario",
        "% Abandono CR", "% Abandono CSR",
        "Espera CR (min)", "Espera CSR (min)",
        "Ociosidad CR (%)", "Ociosidad CSR (%)",
        "Atendidos CR", "Atendidos CSR"
    ]
    for i, h in enumerate(headers_g):
        ref = f"{get_column_letter(i+1)}3"
        cel(ws, ref, h, bold=True, size=9, color=BLANCO, bg=AZUL_MED, border=True)
    set_row_height(ws, 3, 18)

    for idx, r in enumerate(resultados):
        row = 4 + idx
        cr = r["CR"]
        csr = r["CSR"]
        oc_cr = cr["ociosidad_por_cargador"]
        oc_csr = csr["ociosidad_por_cargador"]
        ocio_cr = (sum(oc_cr) / len(oc_cr) / 100) if oc_cr else 0
        ocio_csr = (sum(oc_csr) / len(oc_csr) / 100) if oc_csr else 0
        vals = [
            r["nombre"],
            cr["pct_arrepentidos"] / 100, csr["pct_arrepentidos"] / 100,
            cr["tiempo_espera_prom"], csr["tiempo_espera_prom"],
            ocio_cr, ocio_csr,
            cr["atendidos"], csr["atendidos"]
        ]
        bg = GRIS_CLARO if idx % 2 == 0 else BLANCO
        for i, v in enumerate(vals):
            ref = f"{get_column_letter(i+1)}{row}"
            fmt_map = [None, "0.0%", "0.0%", "0.00", "0.00", "0.0%", "0.0%", "0.0", "0.0"]
            cel(ws, ref, v, size=10, bg=bg, border=True, fmt=fmt_map[i])
        set_row_height(ws, row, 18)

    n = len(resultados)

    # Gráficos comparativos CR vs CSR
    # 1. % Abandono
    ch1 = BarChart()
    ch1.type = "col"
    ch1.title = "% Abandono por Escenario (CR vs CSR)"
    ch1.style = 10; ch1.grouping = "clustered"
    ch1.width = 18; ch1.height = 10
    ch1.y_axis.numFmt = "0%"; ch1.y_axis.title = "Tasa de abandono"
    data1 = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=3 + n)
    cats1 = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    ch1.add_data(data1, titles_from_data=True); ch1.set_categories(cats1)
    ch1.series[0].graphicalProperties.solidFill = AZUL_CLARO  # CR
    ch1.series[1].graphicalProperties.solidFill = "117A65"   # CSR
    ws.add_chart(ch1, "J3")

    # 2. Espera promedio
    ch2 = BarChart()
    ch2.type = "col"
    ch2.title = "Espera Promedio (min) CR vs CSR"
    ch2.style = 10; ch2.width = 18; ch2.height = 10
    ch2.y_axis.title = "Minutos"
    data2 = Reference(ws, min_col=4, min_row=3, max_col=5, max_row=3 + n)
    cats2 = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    ch2.add_data(data2, titles_from_data=True); ch2.set_categories(cats2)
    ch2.series[0].graphicalProperties.solidFill = AZUL_CLARO
    ch2.series[1].graphicalProperties.solidFill = "117A65"
    ws.add_chart(ch2, "J22")

    # 3. Ociosidad
    ch3 = BarChart()
    ch3.type = "col"
    ch3.title = "Ociosidad Promedio (%) CR vs CSR"
    ch3.style = 10; ch3.width = 18; ch3.height = 10
    ch3.y_axis.numFmt = "0%"; ch3.y_axis.title = "Ociosidad"
    data3 = Reference(ws, min_col=6, min_row=3, max_col=7, max_row=3 + n)
    cats3 = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    ch3.add_data(data3, titles_from_data=True); ch3.set_categories(cats3)
    ch3.series[0].graphicalProperties.solidFill = AZUL_CLARO
    ch3.series[1].graphicalProperties.solidFill = "117A65"
    ws.add_chart(ch3, "J41")

    set_col_widths(ws, {
        "A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 2
    })


# ── Hoja 5: Eficiencia global ─────────────────────────────────────────────────
#
# CAMBIO PRINCIPAL: mayor score = mejor (↑).
# - El ranking ahora ordena de MAYOR a MENOR.
# - La descripción de la fórmula se actualiza para reflejar el nuevo sentido.
# - El gráfico de barras muestra las barras más largas como las mejores.
# - Se agrega un segundo gráfico comparando CR vs CSR por escenario
#   (eficiencia desglosada, separada del tipo de cargador).

def _hoja_eficiencia(wb, resultados):
    ws = wb.create_sheet("Eficiencia")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    cel(ws, "A1", "RANKING DE EFICIENCIA GLOBAL",
        bold=True, size=14, color=BLANCO, bg=AZUL_OSCURO)
    set_row_height(ws, 1, 32)

    ws.merge_cells("A2:F2")
    cel(ws, "A2",
        "Score compuesto (↑ mayor = mejor):  "
        "−0.4×Espera − 0.3×Ociosidad − 0.2×Abandono + 0.1×Atendidos",
        italic=True, size=8, color=BLANCO, bg=AZUL_MED, align="left")
    set_row_height(ws, 2, 16)

    hdrs = ["Pos.", "Escenario", "Score", "Evaluación"]
    for i, h in enumerate(hdrs):
        cel(ws, f"{get_column_letter(i+1)}4", h,
            bold=True, size=10, color=BLANCO, bg=AZUL_MED, border=True)
    set_row_height(ws, 4, 20)

    # CAMBIO: ordenar de mayor a menor (mejor primero)
    ranked = sorted(resultados, key=lambda r: r["eficiencia_global"], reverse=True)
    medals = ["🥇", "🥈", "🥉"]

    for rank, r in enumerate(ranked):
        row = 5 + rank
        set_row_height(ws, row, 22)
        bg = [BLANCO, GRIS_CLARO, AZUL_SUAVE][rank % 3]
        score = r["eficiencia_global"]
        eval_txt = "Óptimo" if rank == 0 else ("Aceptable" if rank == 1 else "Mejorable")
        eval_col = [VERDE_OK, NARANJA_WARN, ROJO_BAD][rank]

        cel(ws, f"A{row}", f"{medals[rank]} #{rank+1}",
            bold=True, size=12, bg=bg, border=True)
        cel(ws, f"B{row}", r["nombre"],
            bold=True, size=11, bg=bg, align="left", border=True)
        cel(ws, f"C{row}", score, bold=True, size=12,
            color=AZUL_OSCURO, bg=bg, border=True, fmt="0.000")
        cel(ws, f"D{row}", eval_txt, bold=True, size=11,
            color=eval_col, bg=bg, border=True)

    n = len(resultados)

    # ── Tabla auxiliar para gráfico 1 (ranking) ───────────────────────────
    aux1 = 5 + n + 2
    ws[f"A{aux1}"] = "Escenario"
    ws[f"B{aux1}"] = "Score"
    for i, r in enumerate(ranked):
        ws[f"A{aux1+1+i}"] = r["nombre"]
        ws[f"B{aux1+1+i}"] = r["eficiencia_global"]

    ch1 = BarChart()
    ch1.type = "bar"      # horizontal: la barra más larga = mejor
    ch1.title = "Score de Eficiencia (↑ mayor = mejor)"
    ch1.style = 10; ch1.width = 18; ch1.height = 10
    data1 = Reference(ws, min_col=2, min_row=aux1, max_row=aux1 + n)
    cats1 = Reference(ws, min_col=1, min_row=aux1 + 1, max_row=aux1 + n)
    ch1.add_data(data1, titles_from_data=True)
    ch1.set_categories(cats1)
    ch1.series[0].graphicalProperties.solidFill = AZUL_CLARO
    ws.add_chart(ch1, "F4")

    # ── Tabla auxiliar para gráfico 2: eficiencia desglosada CR vs CSR ────
    # Este gráfico agrupa por escenario y muestra el score de eficiencia
    # para CR y CSR por separado (aclaración: el score global es único por
    # escenario, así que aquí mostramos el % abandono por tipo de cargador
    # como proxy de eficiencia desagregada).

    aux2 = aux1 + n + 3
    ws.merge_cells(f"A{aux2}:D{aux2}")
    cel(ws, f"A{aux2}",
        "Comparación % abandono por tipo de cargador",
        bold=True, size=9, color=BLANCO, bg=AZUL_MED)
    set_row_height(ws, aux2, 18)

    aux2h = aux2 + 1
    for col_idx, header in enumerate(["Escenario", "CR", "CSR"]):
        cel(ws, f"{get_column_letter(col_idx+1)}{aux2h}",
            header, bold=True, size=9, color=BLANCO, bg=AZUL_MED, border=True)
    set_row_height(ws, aux2h, 16)

    for i, r in enumerate(resultados):
        row = aux2h + 1 + i
        val_cr  = r["CR"]["pct_arrepentidos"]  / 100
        val_csr = r["CSR"]["pct_arrepentidos"] / 100
        cel(ws, f"A{row}", r["nombre"], size=9, bg=GRIS_CLARO, border=True)
        cel(ws, f"B{row}", val_cr,  size=9, bg=BLANCO, border=True, fmt="0.0%")
        cel(ws, f"C{row}", val_csr, size=9, bg=BLANCO, border=True, fmt="0.0%")
        set_row_height(ws, row, 16)

    ch2 = BarChart()
    ch2.type = "col"
    ch2.grouping = "clustered"
    ch2.title = "% Abandono por Escenario: CR vs CSR"
    ch2.style = 10; ch2.width = 18; ch2.height = 10
    ch2.y_axis.numFmt = "0%"
    ch2.y_axis.title = "Tasa de abandono"
    data2 = Reference(ws, min_col=2, min_row=aux2h,
                      max_col=3, max_row=aux2h + n)
    cats2 = Reference(ws, min_col=1, min_row=aux2h + 1, max_row=aux2h + n)
    ch2.add_data(data2, titles_from_data=True)
    ch2.set_categories(cats2)
    ch2.series[0].graphicalProperties.solidFill = AZUL_CLARO   # CR
    ch2.series[1].graphicalProperties.solidFill = "117A65"     # CSR (verde)
    ws.add_chart(ch2, "F20")

    set_col_widths(ws, {"A": 10, "B": 18, "C": 12, "D": 14, "E": 2})


# ── Entry point ───────────────────────────────────────────────────────────────

def generar_excel(resultados, path, sim_time, seed, n_replicas):
    wb = Workbook()

    _hoja_resumen(wb, resultados, sim_time, seed, n_replicas)
    _hoja_kpis(wb, resultados)
    _hoja_insights(wb, resultados)
    _hoja_graficos(wb, resultados)
    _hoja_eficiencia(wb, resultados)

    wb.save(path)
    print(f"✓ Excel guardado en: {path}")